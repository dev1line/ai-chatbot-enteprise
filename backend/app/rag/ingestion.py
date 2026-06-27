"""Multimodal ingestion: text/markdown, PDF, Excel/CSV, image (caption/OCR optional).

Each chunk carries citation metadata (doc_id, version, type, page/sheet/cell_range...).
Immutable repository: version + content_hash.
"""
from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass, field
from typing import Any


@dataclass
class Chunk:
    text: str
    metadata: dict[str, Any] = field(default_factory=dict)


def content_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def chunk_text(text: str, chunk_size: int = 800, overlap: int = 120) -> list[str]:
    text = text.strip()
    if not text:
        return []
    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = min(start + chunk_size, len(text))
        chunks.append(text[start:end])
        if end == len(text):
            break
        start = end - overlap
    return chunks


def _base_meta(doc_id: str, version: str, doc_type: str, source: str) -> dict[str, Any]:
    return {"doc_id": doc_id, "version": version, "type": doc_type, "source": source}


def load_text(
    data: bytes, doc_id: str, version: str, source: str, doc_type: str = "text"
) -> list[Chunk]:
    text = data.decode("utf-8", errors="ignore")
    out = []
    for i, piece in enumerate(chunk_text(text)):
        meta = _base_meta(doc_id, version, doc_type, source)
        meta["chunk"] = i
        meta["snippet"] = piece[:160]
        out.append(Chunk(text=piece, metadata=meta))
    return out


def load_pdf(data: bytes, doc_id: str, version: str, source: str) -> list[Chunk]:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    out: list[Chunk] = []
    for page_num, page in enumerate(reader.pages, start=1):
        page_text = (page.extract_text() or "").strip()
        if not page_text:
            continue
        for i, piece in enumerate(chunk_text(page_text)):
            meta = _base_meta(doc_id, version, "pdf", source)
            meta.update({"page": page_num, "chunk": i, "snippet": piece[:160]})
            out.append(Chunk(text=piece, metadata=meta))
    return out


def load_excel(data: bytes, doc_id: str, version: str, source: str) -> list[Chunk]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[Chunk] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else "" for c in rows[0]]
        # Each data row -> a "col=value" description block, keeping the cell reference.
        for r_idx, row in enumerate(rows[1:], start=2):
            parts = []
            for c_idx, val in enumerate(row):
                if val is None:
                    continue
                col_name = header[c_idx] if c_idx < len(header) else f"col{c_idx}"
                parts.append(f"{col_name}={val}")
            if not parts:
                continue
            text = f"[{sheet.title}] " + "; ".join(parts)
            cell_range = f"A{r_idx}:{chr(64 + max(1, len(row)))}{r_idx}"
            meta = _base_meta(doc_id, version, "excel", source)
            meta.update({"sheet": sheet.title, "cell_range": cell_range, "snippet": text[:160]})
            out.append(Chunk(text=text, metadata=meta))
    return out


def load_image(
    data: bytes, doc_id: str, version: str, source: str, caption: str | None = None
) -> list[Chunk]:
    """Image loader: use the caption if available; try OCR (pytesseract) if installed.

    tesseract is not required — gracefully falls back to keep the docker build lightweight.
    """
    text = caption or ""
    try:  # OCR optional
        import pytesseract
        from PIL import Image

        ocr = pytesseract.image_to_string(Image.open(io.BytesIO(data)))
        text = (text + "\n" + ocr).strip()
    except Exception:  # noqa: BLE001
        pass
    if not text:
        text = f"[image] {source}"
    meta = _base_meta(doc_id, version, "image", source)
    meta["snippet"] = text[:160]
    return [Chunk(text=text, metadata=meta)]


def load_document(
    filename: str,
    data: bytes,
    doc_id: str,
    version: str,
    caption: str | None = None,
) -> list[Chunk]:
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return load_pdf(data, doc_id, version, filename)
    if lower.endswith((".xlsx", ".xlsm")):
        return load_excel(data, doc_id, version, filename)
    if lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
        return load_image(data, doc_id, version, filename, caption=caption)
    if lower.endswith(".csv"):
        return load_text(data, doc_id, version, filename, doc_type="excel")
    return load_text(data, doc_id, version, filename, doc_type="text")
